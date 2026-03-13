from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='top', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def _style_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws, min_width=15, max_width=60):
    for col in ws.columns:
        col_letter = col[0].column_letter
        lengths = []
        for cell in col:
            if cell.value:
                lines = str(cell.value).split('\n')
                lengths.append(max(len(line) for line in lines))
        best = max(lengths) + 4 if lengths else min_width
        ws.column_dimensions[col_letter].width = min(max(best, min_width), max_width)


def export_xlsx(ng_results, ok_results, output_path):
    """
    将 NG / OK 结果写入 output.xlsx
    ng_results: list[ParseResult]  状态为 NG 的解析结果
    ok_results: list[ParseResult]  状态为 OK 的解析结果
    """
    wb = Workbook()

    # --- Sheet1: NG ---
    ws_ng = wb.active
    ws_ng.title = 'NG'
    _style_header(ws_ng, ['文件名称', '结果1（Mul Short:）', '结果2（Mutual Short:）'])

    for row_idx, r in enumerate(ng_results, 2):
        name_cell = ws_ng.cell(row=row_idx, column=1, value=r.filename)
        mul_cell = ws_ng.cell(row=row_idx, column=2, value=r.mul_short)
        mut_cell = ws_ng.cell(row=row_idx, column=3, value=r.mutual_short)
        for c in (name_cell, mul_cell, mut_cell):
            c.alignment = CELL_ALIGN
            c.border = THIN_BORDER

    _auto_width(ws_ng)

    # --- Sheet2: OK ---
    ws_ok = wb.create_sheet(title='OK')
    _style_header(ws_ok, ['文件名称'])

    for row_idx, r in enumerate(ok_results, 2):
        name_cell = ws_ok.cell(row=row_idx, column=1, value=r.filename)
        name_cell.alignment = CELL_ALIGN
        name_cell.border = THIN_BORDER

    _auto_width(ws_ok)

    wb.save(output_path)
    return output_path
